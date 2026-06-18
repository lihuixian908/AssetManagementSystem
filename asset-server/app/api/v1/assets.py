from typing import Annotated

from datetime import datetime

from fastapi import APIRouter, Depends, File, Query, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.api.deps import SessionDep, get_current_active_user
from app.models.user import User
from app.models.asset import Asset
from app.schemas.asset import AssetAction, AssetCreate, AssetInDB, AssetUpdate
from app.schemas.common import Response
from app.services import asset_service

router = APIRouter(prefix="/assets", tags=["资产管理"])


@router.get("", response_model=Response)
def get_assets(
    db: SessionDep,
    page: int = Query(1, ge=1),
    page_size: int = Query(10, ge=1, le=100),
    keyword: str | None = None,
    status_filter: str | None = None,
    category_id: int | None = None,
    company_code: str | None = None,
    sn: str | None = None,
):
    result = asset_service.get_asset_list(
        db, page, page_size, keyword, status_filter, category_id, company_code, sn
    )
    return Response(data=result)


@router.get("/export")
def export_assets(db: SessionDep):
    output = asset_service.export_assets_excel(db)
    filename = f"assets_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/import", response_model=Response)
async def import_assets(
    file: UploadFile = File(...),
    db: SessionDep = None,
    current_user: Annotated[User, Depends(get_current_active_user)] = None,
):
    from io import BytesIO
    from openpyxl import load_workbook
    from app.models.category import Category as CatModel
    from app.models.user import User as UserModel
    from sqlalchemy import select
    from fastapi import HTTPException

    if not file.filename or not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .xls 格式")

    # 确保"待分类"存在
    def get_or_create_uncategorized():
        row = db.execute(select(CatModel).where(CatModel.code == "UNCAT")).scalar_one_or_none()
        if not row:
            row = CatModel(name="待分类", code="UNCAT", parent_id=0, sort_order=99)
            db.add(row)
            db.flush()
        return row.id

    uncat_id = get_or_create_uncategorized()

    # 预加载分类和用户映射
    cat_rows = db.execute(select(CatModel.id, CatModel.name)).all()
    cat_map: dict[str, int] = {r[1]: r[0] for r in cat_rows}
    user_rows = db.execute(select(UserModel.id, UserModel.real_name)).all()
    user_map: dict[str, int] = {r[1]: r[0] for r in user_rows}

    content = await file.read()
    wb = load_workbook(BytesIO(content))
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    errors: list[str] = []
    assets_to_add: list[Asset] = []
    seen_codes: set[str] = set()

    for i, row in enumerate(rows, 2):
        if not row or not any(row):
            continue
        vals = [str(c).strip() if c else '' for c in row]
        asset_code = vals[0] if len(vals) > 0 else ''
        name = vals[1] if len(vals) > 1 else ''
        company_code = vals[2] if len(vals) > 2 else ''
        sn_val = vals[3] if len(vals) > 3 else ''
        cat_name = vals[4] if len(vals) > 4 else ''
        dept = vals[5] if len(vals) > 5 else ''
        owner_name = vals[6] if len(vals) > 6 else ''

        if not asset_code:
            errors.append(f"第{i}行: 资产编号不能为空")
            continue
        if not name:
            errors.append(f"第{i}行({asset_code}): 设备名称不能为空")
            continue

        # 查重
        existing = db.execute(select(Asset).where(Asset.asset_code == asset_code)).scalar_one_or_none()
        if existing or asset_code in seen_codes:
            errors.append(f"第{i}行({asset_code}): 资产编号已存在")
            continue
        seen_codes.add(asset_code)

        category_id = cat_map.get(cat_name, uncat_id)
        user_id = user_map.get(owner_name)

        assets_to_add.append(Asset(
            asset_code=asset_code,
            name=name,
            company_code=company_code or None,
            sn=sn_val or None,
            category_id=category_id,
            department=dept or None,
            user_id=user_id,
            status="normal",
        ))

    if assets_to_add:
        db.add_all(assets_to_add)
        db.commit()

    return Response(data={
        "imported": len(assets_to_add),
        "total": len(rows),
        "errors": errors,
    })


@router.get("/{asset_id}", response_model=Response)
def get_asset(asset_id: int, db: SessionDep):
    result = asset_service.get_asset(db, asset_id)
    return Response(data=result)


@router.post("", response_model=Response)
def create_asset(
    asset_in: AssetCreate,
    db: SessionDep,
    current_user: Annotated[User, Depends(get_current_active_user)],
):
    result = asset_service.create_asset(db, asset_in, current_user.real_name)
    return Response(data=result.model_dump())


@router.put("/{asset_id}", response_model=Response)
def update_asset(
    asset_id: int,
    asset_in: AssetUpdate,
    db: SessionDep,
    current_user: Annotated[User, Depends(get_current_active_user)],
):
    result = asset_service.update_asset(db, asset_id, asset_in, current_user.real_name)
    return Response(data=result.model_dump())


@router.delete("/{asset_id}", response_model=Response)
def delete_asset(
    asset_id: int,
    db: SessionDep,
    current_user: Annotated[User, Depends(get_current_active_user)],
):
    asset_service.delete_asset(db, asset_id)
    return Response(message="删除成功")


@router.post("/batch-delete", response_model=Response)
def batch_delete_assets(
    ids: list[int],
    db: SessionDep,
    current_user: Annotated[User, Depends(get_current_active_user)],
):
    from sqlalchemy import select
    assets = db.execute(select(Asset).where(Asset.id.in_(ids))).scalars().all()
    for a in assets:
        db.delete(a)
    db.commit()
    return Response(data={"deleted": len(assets)}, message=f"已删除{len(assets)}台设备")


@router.post("/{asset_id}/assign", response_model=Response)
def assign_asset(
    asset_id: int,
    action: AssetAction,
    user_id: int,
    db: SessionDep,
    current_user: Annotated[User, Depends(get_current_active_user)],
):
    result = asset_service.assign_asset(
        db, asset_id, user_id, current_user.real_name, action.description
    )
    return Response(data=result.model_dump())


@router.post("/{asset_id}/return", response_model=Response)
def return_asset(
    asset_id: int,
    action: AssetAction,
    db: SessionDep,
    current_user: Annotated[User, Depends(get_current_active_user)],
):
    result = asset_service.return_asset(
        db, asset_id, current_user.real_name, action.description
    )
    return Response(data=result.model_dump())


@router.post("/{asset_id}/transfer", response_model=Response)
def transfer_asset(
    asset_id: int,
    action: AssetAction,
    user_id: int | None = None,
    db: SessionDep = None,
    current_user: Annotated[User, Depends(get_current_active_user)] = None,
):
    result = asset_service.transfer_asset(
        db, asset_id, user_id, current_user.real_name, action.description
    )
    return Response(data=result.model_dump())


@router.post("/{asset_id}/maintenance", response_model=Response)
def maintenance_asset(
    asset_id: int,
    action: AssetAction,
    db: SessionDep,
    current_user: Annotated[User, Depends(get_current_active_user)],
):
    result = asset_service.maintenance_asset(
        db, asset_id, current_user.real_name, action.description
    )
    return Response(data=result.model_dump())


@router.post("/{asset_id}/scrap", response_model=Response)
def scrap_asset(
    asset_id: int,
    action: AssetAction,
    db: SessionDep,
    current_user: Annotated[User, Depends(get_current_active_user)],
):
    result = asset_service.scrap_asset(
        db, asset_id, current_user.real_name, action.description
    )
    return Response(data=result.model_dump())


@router.post("/{asset_id}/scan", response_model=Response)
def scan_asset(
    asset_id: int,
    db: SessionDep,
    current_user: Annotated[User, Depends(get_current_active_user)],
):
    result = asset_service.scan_asset(db, asset_id, current_user.real_name)
    return Response(data=result.model_dump())


@router.post("/{asset_id}/borrow", response_model=Response)
def borrow_asset(
    asset_id: int,
    action: AssetAction,
    user_id: int,
    db: SessionDep,
    current_user: Annotated[User, Depends(get_current_active_user)],
):
    result = asset_service.borrow_asset(
        db, asset_id, user_id, current_user.real_name, action.description
    )
    return Response(data=result.model_dump())


@router.post("/{asset_id}/return-borrow", response_model=Response)
def return_borrow_asset(
    asset_id: int,
    action: AssetAction,
    db: SessionDep,
    current_user: Annotated[User, Depends(get_current_active_user)],
):
    result = asset_service.return_borrowed_asset(
        db, asset_id, current_user.real_name, action.description
    )
    return Response(data=result.model_dump())


@router.get("/{asset_id}/records", response_model=Response)
def get_asset_records(
    asset_id: int,
    db: SessionDep,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
):
    result = asset_service.get_asset_records(db, asset_id, page, page_size)
    return Response(data=result)