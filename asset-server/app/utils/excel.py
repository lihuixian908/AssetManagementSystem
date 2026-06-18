import io
from typing import Any

from fastapi import HTTPException


def export_to_excel(data: list[dict[str, Any]], filename: str = "export.xlsx") -> bytes:
    try:
        import openpyxl
        from openpyxl import Workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl 未安装")
    
    wb = Workbook()
    ws = wb.active
    ws.title = filename.split(".")[0]
    
    if not data:
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
    
    headers = list(data[0].keys())
    ws.append(headers)
    
    for item in data:
        ws.append([item.get(h) for h in headers])
    
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def import_from_excel(file_content: bytes) -> list[dict[str, Any]]:
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl 未安装")
    
    wb = openpyxl.load_workbook(filename="data.xlsx", read_only=True, data_only=True)
    ws = wb.active
    
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    
    headers = [str(h) for h in rows[0]]
    data = []
    for row in rows[1:]:
        data.append(dict(zip(headers, row)))
    
    return data