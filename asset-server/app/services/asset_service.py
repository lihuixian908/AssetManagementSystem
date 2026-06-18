from datetime import datetime
from fastapi import HTTPException, status
from sqlalchemy.orm import Session

from app.models.asset import Asset
from app.models.record import AssetRecord
from app.repositories.asset_repo import AssetRepository
from app.schemas.asset import AssetCreate, AssetUpdate, AssetInDB


asset_repo = AssetRepository()

ASSET_STATUS_MAP = {
    "normal": "在库",
    "borrowed": "已借出",
    "scrapped": "已报废",
}

RECORD_TYPE_MAP = {
    "create": "入库",
    "assign": "领用",
    "return": "归还",
    "transfer": "调拨",
    "maintenance": "维修",
    "scrap": "报废",
    "update": "更新",
    "scan": "扫码",
    "borrow": "出借",
}


def _create_record(
    db: Session,
    asset_id: int,
    user_id: int | None,
    record_type: str,
    description: str,
    operator: str,
) -> AssetRecord:
    record = AssetRecord(
        asset_id=asset_id,
        user_id=user_id,
        type=record_type,
        description=description,
        operator=operator,
    )
    db.add(record)
    return record


def create_asset(db: Session, asset_in: AssetCreate, operator: str) -> AssetInDB:
    existing = asset_repo.get_by_asset_code(db, asset_in.asset_code)
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="资产编号已存在",
        )
    asset = asset_repo.create(db, obj_in=asset_in)
    _create_record(
        db,
        asset_id=asset.id,
        user_id=None,
        record_type="create",
        description=f"资产入库：{asset.name}",
        operator=operator,
    )
    db.commit()
    return AssetInDB.model_validate(asset)


def get_asset(db: Session, asset_id: int) -> dict:
    asset = asset_repo.get(db, asset_id)
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="资产不存在",
        )
    result = AssetInDB.model_validate(asset).model_dump()
    if asset.user_id:
        from app.models.user import User as UserModel
        user = db.get(UserModel, asset.user_id)
        result["owner_name"] = user.real_name if user else None
    else:
        result["owner_name"] = None
    return result


def get_asset_list(
    db: Session,
    page: int = 1,
    page_size: int = 10,
    keyword: str | None = None,
    status_filter: str | None = None,
    category_id: int | None = None,
    company_code: str | None = None,
    sn: str | None = None,
) -> dict:
    from sqlalchemy import or_, select, func, case
    
    skip = (page - 1) * page_size
    query = select(Asset)
    count_query = select(func.count(Asset.id))
    
    if keyword:
        kw = f"%{keyword}%"
        query = query.where(
            or_(
                Asset.asset_code.like(kw),
                Asset.name.like(kw),
                Asset.location.like(kw),
                Asset.description.like(kw),
            )
        )
        count_query = count_query.where(
            or_(
                Asset.asset_code.like(kw),
                Asset.name.like(kw),
                Asset.location.like(kw),
                Asset.description.like(kw),
            )
        )
    
    if status_filter:
        query = query.where(Asset.status == status_filter)
        count_query = count_query.where(Asset.status == status_filter)
    
    if category_id:
        query = query.where(Asset.category_id == category_id)
        count_query = count_query.where(Asset.category_id == category_id)

    if company_code:
        query = query.where(Asset.company_code.ilike(f"%{company_code}%"))
        count_query = count_query.where(Asset.company_code.ilike(f"%{company_code}%"))

    if sn:
        query = query.where(Asset.sn.ilike(f"%{sn}%"))
        count_query = count_query.where(Asset.sn.ilike(f"%{sn}%"))

    total = db.execute(count_query).scalar_one()

    status_order = case(
        (Asset.status == "normal", 1),
        (Asset.status == "borrowed", 2),
        (Asset.status == "scrapped", 3),
        else_=4,
    )
    query = query.order_by(status_order, Asset.created_at.desc()).offset(skip).limit(page_size)
    assets = list(db.execute(query).scalars().all())

    # 批量查用户姓名
    from app.models.user import User as UserModel
    user_ids = list({a.user_id for a in assets if a.user_id})
    user_map: dict[int, str] = {}
    if user_ids:
        user_query = select(UserModel.id, UserModel.real_name).where(UserModel.id.in_(user_ids))
        for row in db.execute(user_query).all():
            user_map[row[0]] = row[1]

    items = []
    for a in assets:
        item = AssetInDB.model_validate(a).model_dump()
        item["owner_name"] = user_map.get(a.user_id) if a.user_id else None
        items.append(item)

    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "items": items,
    }


def update_asset(
    db: Session, asset_id: int, asset_in: AssetUpdate, operator: str
) -> AssetInDB:
    asset = asset_repo.get(db, asset_id)
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="资产不存在",
        )
    asset = asset_repo.update(db, db_obj=asset, obj_in=asset_in)
    _create_record(
        db,
        asset_id=asset.id,
        user_id=None,
        record_type="update",
        description=f"资产信息更新：{asset.name}",
        operator=operator,
    )
    db.commit()
    return AssetInDB.model_validate(asset)


def delete_asset(db: Session, asset_id: int) -> bool:
    return asset_repo.delete(db, id=asset_id)


def assign_asset(
    db: Session, asset_id: int, user_id: int, operator: str, description: str | None = None
) -> AssetInDB:
    asset = asset_repo.get(db, asset_id)
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="资产不存在",
        )
    if asset.status == "scrapped":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="资产已报废，不可领用",
        )
    asset.user_id = user_id
    asset.status = "normal"
    db.add(asset)
    _create_record(
        db,
        asset_id=asset.id,
        user_id=user_id,
        record_type="assign",
        description=description or f"资产领用：{asset.name}",
        operator=operator,
    )
    db.commit()
    db.refresh(asset)
    return AssetInDB.model_validate(asset)


def return_asset(
    db: Session, asset_id: int, operator: str, description: str | None = None
) -> AssetInDB:
    asset = asset_repo.get(db, asset_id)
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="资产不存在",
        )
    if asset.status != "normal":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="资产当前状态不可归还",
        )
    asset.user_id = None
    db.add(asset)
    _create_record(
        db,
        asset_id=asset.id,
        user_id=None,
        record_type="return",
        description=description or f"资产归还：{asset.name}",
        operator=operator,
    )
    db.commit()
    db.refresh(asset)
    return AssetInDB.model_validate(asset)


def transfer_asset(
    db: Session,
    asset_id: int,
    user_id: int | None,
    operator: str,
    description: str | None = None,
) -> AssetInDB:
    asset = asset_repo.get(db, asset_id)
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="资产不存在",
        )
    asset.user_id = user_id
    db.add(asset)
    _create_record(
        db,
        asset_id=asset.id,
        user_id=None,
        record_type="transfer",
        description=description or f"资产调拨：{asset.name}",
        operator=operator,
    )
    db.commit()
    db.refresh(asset)
    return AssetInDB.model_validate(asset)


def maintenance_asset(
    db: Session, asset_id: int, operator: str, description: str | None = None
) -> AssetInDB:
    asset = asset_repo.get(db, asset_id)
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="资产不存在",
        )
    if asset.status == "scrapped":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="资产已报废，不可维修",
        )
    _create_record(
        db,
        asset_id=asset.id,
        user_id=None,
        record_type="maintenance",
        description=description or f"资产维修：{asset.name}",
        operator=operator,
    )
    db.commit()
    db.refresh(asset)
    return AssetInDB.model_validate(asset)


def scrap_asset(
    db: Session, asset_id: int, operator: str, description: str | None = None
) -> AssetInDB:
    asset = asset_repo.get(db, asset_id)
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="资产不存在",
        )
    asset.status = "scrapped"
    db.add(asset)
    _create_record(
        db,
        asset_id=asset.id,
        user_id=None,
        record_type="scrap",
        description=description or f"资产报废：{asset.name}",
        operator=operator,
    )
    db.commit()
    db.refresh(asset)
    return AssetInDB.model_validate(asset)


def borrow_asset(
    db: Session, asset_id: int, user_id: int, operator: str, description: str | None = None
) -> AssetInDB:
    asset = asset_repo.get(db, asset_id)
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="资产不存在",
        )
    if asset.status == "scrapped":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="资产已报废，不可出借",
        )
    asset.user_id = user_id
    _create_record(
        db,
        asset_id=asset.id,
        user_id=user_id,
        record_type="borrow",
        description=description or f"资产出借：{asset.name}",
        operator=operator,
    )
    db.commit()
    db.refresh(asset)
    return AssetInDB.model_validate(asset)


def return_borrowed_asset(
    db: Session, asset_id: int, operator: str, description: str | None = None
) -> AssetInDB:
    asset = asset_repo.get(db, asset_id)
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="资产不存在",
        )
    asset.user_id = None
    _create_record(
        db,
        asset_id=asset.id,
        user_id=None,
        record_type="return",
        description=description or f"出借资产归还：{asset.name}",
        operator=operator,
    )
    db.commit()
    db.refresh(asset)
    return AssetInDB.model_validate(asset)


def scan_asset(
    db: Session, asset_id: int, operator: str, description: str | None = None
) -> AssetInDB:
    asset = asset_repo.get(db, asset_id)
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="资产不存在",
        )
    _create_record(
        db,
        asset_id=asset.id,
        user_id=None,
        record_type="scan",
        description=description or f"扫码查看资产：{asset.name}",
        operator=operator,
    )
    db.commit()
    return AssetInDB.model_validate(asset)


def get_asset_records(db: Session, asset_id: int, page: int = 1, page_size: int = 20) -> dict:
    from sqlalchemy import select, func
    
    skip = (page - 1) * page_size
    query = select(AssetRecord).where(AssetRecord.asset_id == asset_id)
    count_query = select(func.count(AssetRecord.id)).where(AssetRecord.asset_id == asset_id)
    
    total = db.execute(count_query).scalar_one()
    query = query.order_by(AssetRecord.created_at.desc()).offset(skip).limit(page_size)
    records = list(db.execute(query).scalars().all())
    
    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "items": [
            {
                "id": r.id,
                "asset_id": r.asset_id,
                "user_id": r.user_id,
                "type": r.type,
                "type_name": RECORD_TYPE_MAP.get(r.type, r.type),
                "description": r.description,
                "operator": r.operator,
                "created_at": r.created_at,
            }
            for r in records
        ],
    }


def export_assets_excel(db: Session):
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from app.models.user import User as UserModel
    from app.models.category import Category

    wb = Workbook()
    ws = wb.active
    ws.title = "资产列表"

    headers = [
        "资产编号", "设备名称", "分类", "公司编号", "SN号",
        "状态", "负责人", "所属部门", "存放位置", "采购日期",
        "备注", "创建时间", "更新时间",
    ]
    header_font = Font(bold=True, size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # 查询全部资产
    from sqlalchemy import select
    query = select(Asset).order_by(Asset.created_at.desc())
    assets = list(db.execute(query).scalars().all())

    # 查用户和分类映射
    user_ids = list({a.user_id for a in assets if a.user_id})
    user_map = {}
    if user_ids:
        for row in db.execute(select(UserModel.id, UserModel.real_name).where(UserModel.id.in_(user_ids))).all():
            user_map[row[0]] = row[1]

    cat_ids = list({a.category_id for a in assets})
    cat_map = {}
    if cat_ids:
        for row in db.execute(select(Category.id, Category.name).where(Category.id.in_(cat_ids))).all():
            cat_map[row[0]] = row[1]

    for row_idx, a in enumerate(assets, 2):
        ws.cell(row=row_idx, column=1, value=a.asset_code)
        ws.cell(row=row_idx, column=2, value=a.name)
        ws.cell(row=row_idx, column=3, value=cat_map.get(a.category_id, ""))
        ws.cell(row=row_idx, column=4, value=a.company_code or "")
        ws.cell(row=row_idx, column=5, value=a.sn or "")
        ws.cell(row=row_idx, column=6, value=ASSET_STATUS_MAP.get(a.status, a.status))
        ws.cell(row=row_idx, column=7, value=user_map.get(a.user_id, "") if a.user_id else "")
        ws.cell(row=row_idx, column=8, value=a.department or "")
        ws.cell(row=row_idx, column=9, value=a.location or "")
        ws.cell(row=row_idx, column=10, value=str(a.purchase_date) if a.purchase_date else "")
        ws.cell(row=row_idx, column=11, value=a.description or "")
        ws.cell(row=row_idx, column=12, value=str(a.created_at)[:19] if a.created_at else "")
        ws.cell(row=row_idx, column=13, value=str(a.updated_at)[:19] if a.updated_at else "")

    # 自动列宽
    for col in range(1, len(headers) + 1):
        max_len = len(headers[col - 1])
        for row in range(2, len(assets) + 2):
            val = ws.cell(row=row, column=col).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = min(max_len + 4, 40)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output